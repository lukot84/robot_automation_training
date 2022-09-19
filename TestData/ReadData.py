import openpyxl

wk = openpyxl.load_workbook("C://Users/Lukasz_Kot/Documents/pyxlTest.xlsx")

def fetch_number_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row

def fetch_cell_data(sheetname, row, cell):
    sh = wk[sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value

print(fetch_number_of_rows('Sheet1'))
print(fetch_cell_data('Sheet1', 1 , 1))