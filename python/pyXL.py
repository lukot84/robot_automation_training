import openpyxl

wk = openpyxl.Workbook()
sh = wk.active
sh.title = "NewSheet"
sh['A4'].value = "tyle"

wk.save("C:\\Users\Lukasz_Kot\Documents\pyxlTest1.xlsx")

# rows = sh.max_row
# columns = sh.max_column

# for i in range(1, rows+1):
#     for j in range(1, columns+1):
#         c = sh.cell(i,j)
#         print(c.value)

